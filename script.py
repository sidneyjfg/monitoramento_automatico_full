import os
import subprocess
import openpyxl
import csv
import re

# Função para extrair letras maiúsculas do nome do arquivo
def extrair_maiusculas(nome_arquivo):
    return ''.join(re.findall(r'[A-Z]', nome_arquivo))

# Função para formatar valores numéricos corretamente
def formatar_valor(valor):
    if isinstance(valor, float):
        if valor.is_integer():  # Se o número é inteiro, remover ".0"
            return str(int(valor))
    return str(valor)

# Função para converter a segunda aba de um arquivo xlsx para csv
def xlsx_to_csv(xlsx_file, output_dir, delimiter='#'):
    try:
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        if len(wb.sheetnames) < 2:
            print(f"O arquivo {xlsx_file} não possui uma segunda aba.")
            return None
        
        sheet = wb[wb.sheetnames[1]]
        nome_base = "notas"
        maiusculas = extrair_maiusculas(os.path.basename(xlsx_file))

        # Concatenar as letras maiúsculas, se houver
        if maiusculas:
            nome_base += maiusculas

        csv_filename = f"{nome_base}.csv"
        csv_file = os.path.join(output_dir, csv_filename)
        
        with open(csv_file, mode='w', newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in sheet.iter_rows(values_only=True):
                formatted_row = [formatar_valor(cell) for cell in row]
                writer.writerow(formatted_row)
        
        print(f"Convertido: {xlsx_file} -> {csv_file}")
        return csv_file
    except Exception as e:
        print(f"Erro ao processar {xlsx_file}: {e}")
        return None

# Função para enviar arquivo via SCP e rodar o script remoto
def enviar_arquivo(cliente_nome, arquivos):
    cliente_user = os.getenv(f"CLIENTE_{cliente_nome.upper()}_USER")
    cliente_server = os.getenv(f"CLIENTE_{cliente_nome.upper()}_SERVER")
    cliente_port = os.getenv(f"CLIENTE_{cliente_nome.upper()}_PORT")
    cliente_password = os.getenv(f"CLIENTE_{cliente_nome.upper()}_PASSWORD")
    cliente_path = os.getenv(f"CLIENTE_{cliente_nome.upper()}_PATH")

    # Exibir a senha capturada para verificar como está sendo recebida
    print(f"Senha capturada para {cliente_nome}: {cliente_password}")

    if not cliente_user or not cliente_server:
        print(f"Credenciais do cliente {cliente_nome} não encontradas.")
        return

    for arquivo in arquivos:
        if not os.path.isfile(arquivo):
            print(f"Arquivo {arquivo} não encontrado para {cliente_nome}.")
            continue

        # Ajuste o comando SCP envolvendo a senha com aspas simples e escape correto
        comando_scp = f"sshpass -p '{cliente_password}' scp -o StrictHostKeyChecking=no -P{cliente_port} {arquivo} {cliente_user}@{cliente_server}:{cliente_path}"
        print(f"Comando SCP: {comando_scp}")
        subprocess.run(comando_scp, shell=True, capture_output=False, check=True)
    
    executar_importa_notas = os.getenv("EXECUTAR_IMPORTA_NOTAS", "false").lower() == "true"
    
    if executar_importa_notas:
        # Executa o script remoto apenas se a variável de controle permitir
        comando_ssh = (
            f"sshpass -p '{cliente_password}' ssh -o StrictHostKeyChecking=no -p{cliente_port} {cliente_user}@{cliente_server} "
            f"'sudo su - eacadm -c \"/u/saci/shells/importa_notas.sh\" 2>&1'"
        )
        print(f"Comando SSH: {comando_ssh}")
        subprocess.run(comando_ssh, shell=True, capture_output=False, check=True)
    else:
        print(f"Execução do script remoto 'importa_notas' desativada para o cliente {cliente_nome}.")




# Função para buscar dinamicamente os clientes nas variáveis de ambiente
def buscar_clientes():
    clientes = []
    for key, value in os.environ.items():
        if key.startswith("CLIENTE_") and key.endswith("_USER"):
            cliente_nome = key.split("_")[1].lower()  # Extrai o nome do cliente a partir da variável
            clientes.append(cliente_nome)
    return clientes

# Função principal para processar os clientes e enviar arquivos
def processar_clientes():
    # Pega o diretório base a partir da variável de ambiente BASE_DIR
    downloads_dir = os.getenv('BASE_DIR', '/home/')  # Diretório base definido no docker-compose.yml

    # Buscar dinamicamente os clientes a partir das variáveis de ambiente
    clientes = buscar_clientes()

    for cliente_nome in clientes:
        arquivos_especiais = os.getenv(f"CLIENTE_{cliente_nome.upper()}_ARQUIVOS_ESPECIAIS")
        
        if arquivos_especiais:
            sufixos = arquivos_especiais.split(',')
            arquivos_excel = [os.path.join(downloads_dir, f"{cliente_nome}{sufixo}.xlsx") for sufixo in sufixos]
        else:
            arquivos_excel = [os.path.join(downloads_dir, f"{cliente_nome}.xlsx")]

        arquivos_csv = []
        
        for xlsx_file in arquivos_excel:
            if os.path.isfile(xlsx_file):
                # Verificar se a pasta de saída para o cliente existe
                output_dir = os.path.join(downloads_dir, cliente_nome)
                
                if not os.path.exists(output_dir):
                    # Criar a pasta do cliente se não existir
                    os.makedirs(output_dir)
                    print(f"Diretório criado para o cliente {cliente_nome}: {output_dir}")

                # Converter o arquivo Excel para CSV
                csv_file = xlsx_to_csv(xlsx_file, output_dir)
                if csv_file:
                    arquivos_csv.append(csv_file)
            else:
                print(f"Arquivo {xlsx_file} não encontrado para {cliente_nome}.")
        
        # Enviar os arquivos convertidos
        if arquivos_csv:
            enviar_arquivo(cliente_nome, arquivos_csv)

# Executar o processo
if __name__ == "__main__":
    processar_clientes()
    print("Envio finalizado com sucesso!")
